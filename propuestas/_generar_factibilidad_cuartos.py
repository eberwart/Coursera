#!/usr/bin/env python3
"""Genera la planilla de factibilidad: cuartos de café para Tostado Club."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUT = str(Path(__file__).resolve().parent / "factibilidad-cuartos-tostado-club.xlsx")

# Paleta
CAFE = "3C2415"
CAFE_MED = "6B3F2A"
DORADO = "C4A35A"
CREMA = "FBF6F0"
CREMA2 = "F3EDE3"
INPUT = "FFF3CD"
INPUT_BORDE = "C4A35A"
BLANCO = "FFFFFF"
VERDE = "1F7A4D"
VERDE_BG = "D8F3E2"
ROJO = "9B2335"
ROJO_BG = "F8D7DA"
GRIS = "5C534A"
GRIS_CLARO = "EFEAE3"
AZUL = "2C5F6E"
AZUL_BG = "D6EAF0"

thin = Border(
    left=Side(style="thin", color="D4C4B0"),
    right=Side(style="thin", color="D4C4B0"),
    top=Side(style="thin", color="D4C4B0"),
    bottom=Side(style="thin", color="D4C4B0"),
)
thick_gold = Border(
    left=Side(style="medium", color=DORADO),
    right=Side(style="medium", color=DORADO),
    top=Side(style="medium", color=DORADO),
    bottom=Side(style="medium", color=DORADO),
)

font_title = Font(name="Calibri", size=20, bold=True, color=BLANCO)
font_sub = Font(name="Calibri", size=12, italic=True, color=CREMA)
font_h = Font(name="Calibri", size=12, bold=True, color=BLANCO)
font_h2 = Font(name="Calibri", size=11, bold=True, color=CAFE)
font_label = Font(name="Calibri", size=11, color=CAFE)
font_input = Font(name="Calibri", size=12, bold=True, color=CAFE)
font_num = Font(name="Calibri", size=11, color="1A1A1A")
font_kpi = Font(name="Calibri", size=18, bold=True, color=CAFE)
font_kpi_w = Font(name="Calibri", size=18, bold=True, color=BLANCO)
font_small = Font(name="Calibri", size=9, italic=True, color=GRIS)
font_white = Font(name="Calibri", size=11, color=BLANCO)
font_white_b = Font(name="Calibri", size=11, bold=True, color=BLANCO)

fill_cafe = PatternFill("solid", fgColor=CAFE)
fill_med = PatternFill("solid", fgColor=CAFE_MED)
fill_gold = PatternFill("solid", fgColor=DORADO)
fill_crema = PatternFill("solid", fgColor=CREMA)
fill_crema2 = PatternFill("solid", fgColor=CREMA2)
fill_input = PatternFill("solid", fgColor=INPUT)
fill_white = PatternFill("solid", fgColor=BLANCO)
fill_verde = PatternFill("solid", fgColor=VERDE_BG)
fill_rojo = PatternFill("solid", fgColor=ROJO_BG)
fill_gris = PatternFill("solid", fgColor=GRIS_CLARO)
fill_azul = PatternFill("solid", fgColor=AZUL_BG)
fill_verde_s = PatternFill("solid", fgColor=VERDE)
fill_rojo_s = PatternFill("solid", fgColor=ROJO)

CLP = '"$"#,##0'
CLP_D = '"$"#,##0.00'
PCT = "0.0%"
PCT2 = "0.00%"
N0 = "#,##0"
N1 = "#,##0.0"
N2 = "#,##0.00"

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")


def fill_row(ws, row, cols, fill):
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = fill


def header_bar(ws, row, cols, title, subtitle=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, title)
    c.font = font_title
    c.fill = fill_cafe
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    fill_row(ws, row, cols, fill_cafe)
    ws.row_dimensions[row].height = 32
    if subtitle:
        r2 = row + 1
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=cols)
        s = ws.cell(r2, 1, subtitle)
        s.font = font_sub
        s.fill = fill_med
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        fill_row(ws, r2, cols, fill_med)
        ws.row_dimensions[r2].height = 22


def section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.font = font_h
    c.fill = fill_med
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    fill_row(ws, row, cols, fill_med)
    ws.row_dimensions[row].height = 20


def style_input(cell, fmt=None):
    cell.fill = fill_input
    cell.font = font_input
    cell.border = thick_gold
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt


def style_out(cell, fmt=None, fill=None):
    cell.font = font_num
    cell.border = thin
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = fill or fill_white
    if fmt:
        cell.number_format = fmt


def label(ws, row, col, text, fill=None):
    c = ws.cell(row, col, text)
    c.font = font_label
    c.alignment = left
    c.fill = fill or fill_crema
    c.border = thin
    return c


def note(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.font = font_small
    c.alignment = left
    return c


def apply_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = Workbook()

    # ------------------------------------------------------------------
    # HOJA PARÁMETROS  (se crea primero para que las refs existan)
    # ------------------------------------------------------------------
    ws_p = wb.active
    ws_p.title = "Parametros"

    COLS_P = 6
    apply_widths(ws_p, [38, 18, 16, 62, 18, 18])
    header_bar(
        ws_p,
        1,
        COLS_P,
        "COFFEE ROASTING LABS  ·  Tostaduría Erik Berwart Araya EIRL",
        "Factibilidad de cuartos (250 g) para Tostado Club  ·  Lote de 1 tonelada de café verde",
    )

    ws_p.merge_cells("A3:F3")
    hint = ws_p["A3"]
    hint.value = (
        "Celdas amarillas = se editan. El resto de la planilla se recalcula solo. "
        "Todos los precios se ingresan NETOS (sin IVA). El IVA se aplica después."
    )
    hint.font = Font(name="Calibri", size=11, italic=True, color=CAFE)
    hint.fill = fill_gold
    hint.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    fill_row(ws_p, 3, COLS_P, fill_gold)
    ws_p.row_dimensions[3].height = 28

    # Encabezados de columnas
    for col, txt in enumerate(["Concepto", "Valor (editar)", "Unidad", "Notas / fuente del supuesto", "", ""], 1):
        c = ws_p.cell(4, col, txt)
        c.font = font_white_b
        c.fill = fill_cafe
        c.alignment = center
        c.border = thin

    # --- Volumen y proceso ---
    section(ws_p, 5, COLS_P, "1. Volumen y proceso")

    rows_vol = [
        (6, "Kg café verde a comprar", 1000, "kg", N0,
         "Lote pedido: 1 tonelada. Cambia este número si el pedido es otro."),
        (7, "Merma de tostado", 0.18, "%", PCT,
         "Pérdida de peso al tostar (humedad + cascarilla). Tueste medio ≈ 16–18%; oscuro ≈ 18–20%. "
         "De 1 kg verde quedan ~820 g tostados con 18%."),
        (8, "Peso del cuarto", 250, "g", N0,
         "Bolsa de 250 g de café TOSTADO (no verde)."),
        (9, "Merma de empaque extra", 0.02, "%", PCT,
         "Bolsas/etiquetas de más por rotura, prueba o error. 0% si compras exacto."),
    ]
    for r, lab, val, uni, fmt, nota in rows_vol:
        label(ws_p, r, 1, lab)
        ws_p.cell(r, 2, val)
        style_input(ws_p.cell(r, 2), fmt)
        ws_p.cell(r, 3, uni).font = font_small
        ws_p.cell(r, 3).alignment = center
        ws_p.cell(r, 3).fill = fill_crema
        ws_p.cell(r, 3).border = thin
        note(ws_p, r, 4, nota)
        ws_p.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws_p.row_dimensions[r].height = 32

    # --- Precios ---
    section(ws_p, 10, COLS_P, "2. Precios (ingresar NETOS, sin IVA)")

    rows_pre = [
        (11, "IVA", 0.19, "%", PCT, "IVA Chile vigente. Afecta caja (crédito/débito), no el margen neto."),
        (12, "Precio café verde (caso base)", 9990, "$/kg neto", CLP,
         "Dato del negocio: $9.990 + IVA / kg. Los escenarios de abajo se comparan contra este."),
        (13, "Maquila (tostado)", 1500, "$/kg neto", CLP,
         "Servicio de tostado: $1.500 + IVA / kg. Ver celda siguiente: ¿se cobra sobre verde o tostado?"),
        (14, "Maquila se cobra sobre", "Verde", "", None,
         "Verde = se multiplica por los kg comprados. Tostado = se multiplica por los kg que salen del tostador."),
        (15, "Bolsa 250 g (válvula / doypack)", 300, "$/un neto", CLP,
         "Supuesto editable. Referencias Chile 2026: Funsmart ≈ $286 neto/un (pack 50); "
         "Nuestro Café ≈ $600; Cafestore ≈ $700. A granel puede bajar. Pon 0 si las pone Tostado Club."),
        (16, "Etiqueta adhesiva", 80, "$/un neto", CLP,
         "Supuesto editable. Referencia imprenta ~$65–$160 + IVA c/u en tirajes de 3.000–4.000. Pon 0 si las pone el cliente."),
        (17, "Precio de venta al cliente (cuarto)", 6000, "$/un neto", CLP,
         "Lo que paga Tostado Club: $6.000 + IVA por bolsa de 250 g."),
    ]
    for r, lab, val, uni, fmt, nota in rows_pre:
        label(ws_p, r, 1, lab)
        ws_p.cell(r, 2, val)
        if fmt:
            style_input(ws_p.cell(r, 2), fmt)
        else:
            style_input(ws_p.cell(r, 2))
        ws_p.cell(r, 3, uni).font = font_small
        ws_p.cell(r, 3).alignment = center
        ws_p.cell(r, 3).fill = fill_crema
        ws_p.cell(r, 3).border = thin
        note(ws_p, r, 4, nota)
        ws_p.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws_p.row_dimensions[r].height = 36

    dv = DataValidation(type="list", formula1='"Verde,Tostado"', allow_blank=False)
    dv.error = "Elige Verde o Tostado"
    dv.errorTitle = "Valor no válido"
    dv.prompt = "Verde o Tostado"
    dv.add("B14")
    ws_p.add_data_validation(dv)

    # --- Costos opcionales ---
    section(ws_p, 18, COLS_P, "3. Costos opcionales del lote (neto) — déjalos en 0 si no aplican")

    rows_op = [
        (19, "Flete / logística del lote", 0, "$ lote neto", CLP,
         "Traslado del verde, del tostado o entrega a Tostado Club. Completar cuando tengas cotización."),
        (20, "Mano de obra de empaque", 0, "$/cuarto neto", CLP,
         "Si empacas tú: costo de personal por bolsa. 0 si está cubierto o no lo quieres costear aún."),
        (21, "Otros costos del lote", 0, "$ lote neto", CLP,
         "Cinta, cajas master, molienda extra, control de calidad, merma de calidad, etc."),
    ]
    for r, lab, val, uni, fmt, nota in rows_op:
        label(ws_p, r, 1, lab)
        ws_p.cell(r, 2, val)
        style_input(ws_p.cell(r, 2), fmt)
        ws_p.cell(r, 3, uni).font = font_small
        ws_p.cell(r, 3).alignment = center
        ws_p.cell(r, 3).fill = fill_crema
        ws_p.cell(r, 3).border = thin
        note(ws_p, r, 4, nota)
        ws_p.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws_p.row_dimensions[r].height = 30

    # --- Escenarios ---
    section(ws_p, 22, COLS_P, "4. Escenarios de precio del café verde (neto, $/kg) — hoja Escenarios")

    escenarios = [
        (23, "Escenario A", 9000, "Pedido: ver qué pasa a $9.000 + IVA"),
        (24, "Escenario B", 9500, "Pedido: ver qué pasa a $9.500 + IVA"),
        (25, "Escenario C (base)", 9990, "Precio actual cotizado: $9.990 + IVA"),
        (26, "Escenario D", 10000, "Pedido: ver qué pasa a $10.000 + IVA"),
        (27, "Escenario E", 10500, "Pedido: ver qué pasa a $10.500 + IVA"),
    ]
    for r, lab, val, nota in escenarios:
        label(ws_p, r, 1, lab)
        ws_p.cell(r, 2, val)
        style_input(ws_p.cell(r, 2), CLP)
        ws_p.cell(r, 3, "$/kg neto").font = font_small
        ws_p.cell(r, 3).alignment = center
        ws_p.cell(r, 3).fill = fill_crema
        ws_p.cell(r, 3).border = thin
        note(ws_p, r, 4, nota)
        ws_p.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    # --- Criterio ---
    section(ws_p, 28, COLS_P, "5. Criterio de factibilidad")

    label(ws_p, 29, 1, "Margen neto mínimo aceptable")
    ws_p["B29"] = 0.25
    style_input(ws_p["B29"], PCT)
    ws_p["C29"] = "%"
    ws_p["C29"].font = font_small
    ws_p["C29"].alignment = center
    ws_p["C29"].fill = fill_crema
    ws_p["C29"].border = thin
    note(ws_p, 29, 4, "Si el margen del lote queda bajo este %, el Resumen marca NO FACTIBLE. Ajústalo a tu criterio.")
    ws_p.merge_cells("D29:F29")

    names = {
        "KgVerde": "Parametros!$B$6",
        "MermaTostado": "Parametros!$B$7",
        "PesoCuartoG": "Parametros!$B$8",
        "MermaEmpaque": "Parametros!$B$9",
        "IVA": "Parametros!$B$11",
        "PrecioVerdeBase": "Parametros!$B$12",
        "MaquilaKg": "Parametros!$B$13",
        "MaquilaSobre": "Parametros!$B$14",
        "PrecioBolsa": "Parametros!$B$15",
        "PrecioEtiqueta": "Parametros!$B$16",
        "PrecioVentaCuarto": "Parametros!$B$17",
        "FleteLote": "Parametros!$B$19",
        "MOEmpaqueCuarto": "Parametros!$B$20",
        "OtrosLote": "Parametros!$B$21",
        "EscA": "Parametros!$B$23",
        "EscB": "Parametros!$B$24",
        "EscC": "Parametros!$B$25",
        "EscD": "Parametros!$B$26",
        "EscE": "Parametros!$B$27",
        "MargenObjetivo": "Parametros!$B$29",
    }
    for n, ref in names.items():
        wb.defined_names.add(DefinedName(name=n, attr_text=ref))

    ws_p.freeze_panes = "A5"
    ws_p.sheet_properties.tabColor = DORADO
    ws_p.page_setup.orientation = "landscape"
    ws_p.page_setup.fitToPage = True
    ws_p.page_setup.fitToWidth = 1
    ws_p.page_setup.fitToHeight = 1
    ws_p.print_title_rows = "1:2"
    ws_p.sheet_view.showGridLines = False
    ws_p.oddHeader.left.text = "Coffee Roasting Labs"
    ws_p.oddFooter.left.text = "Factibilidad cuartos Tostado Club — celdas amarillas editables"

    # ------------------------------------------------------------------
    # HOJA CALC (oculta) — cálculos del caso base
    # ------------------------------------------------------------------
    ws_c = wb.create_sheet("Calc")
    # B column values
    # Row map:
    # 2 KgTostado
    # 3 KgPorCuarto
    # 4 Cuartos
    # 5 SobranteG
    # 6 Bolsas
    # 7 KgMaquila
    formulas = {
        2: ("KgTostado", "=KgVerde*(1-MermaTostado)"),
        3: ("KgPorCuarto", "=PesoCuartoG/1000"),
        4: ("Cuartos", "=IF(B3=0,0,INT(B2/B3))"),
        5: ("SobranteG", "=(B2-B4*B3)*1000"),
        6: ("Bolsas", "=ROUNDUP(B4*(1+MermaEmpaque),0)"),
        7: ("KgMaquila", '=IF(MaquilaSobre="Tostado",B2,KgVerde)'),
        9: ("CostoVerde", "=KgVerde*PrecioVerdeBase"),
        10: ("CostoMaquila", "=B7*MaquilaKg"),
        11: ("CostoBolsas", "=B6*PrecioBolsa"),
        12: ("CostoEtiquetas", "=B6*PrecioEtiqueta"),
        13: ("CostoMO", "=B4*MOEmpaqueCuarto"),
        14: ("CostoFlete", "=FleteLote"),
        15: ("CostoOtros", "=OtrosLote"),
        16: ("CostoTotal", "=B9+B10+B11+B12+B13+B14+B15"),
        17: ("CostoSinVerde", "=B16-B9"),
        19: ("VentaNeta", "=B4*PrecioVentaCuarto"),
        20: ("Utilidad", "=B19-B16"),
        21: ("Margen", "=IF(B19=0,0,B20/B19)"),
        22: ("Markup", "=IF(B16=0,0,B20/B16)"),
        24: ("CostoCuarto", "=IF(B4=0,0,B16/B4)"),
        25: ("CostoKgTostado", "=IF(B2=0,0,B16/B2)"),
        26: ("UtilidadCuarto", "=PrecioVentaCuarto-B24"),
        27: ("PrecioKgTostadoVenta", "=IF(B3=0,0,PrecioVentaCuarto/B3)"),
        29: ("IvaCredito", "=B16*IVA"),
        30: ("IvaDebito", "=B19*IVA"),
        31: ("IvaPagar", "=B30-B29"),
        32: ("CajaSalida", "=B16*(1+IVA)"),
        33: ("CajaEntrada", "=B19*(1+IVA)"),
        34: ("FlujoCaja", "=B33-B32"),
        36: ("PrecioVerdeMaxCero", "=IF(KgVerde=0,0,(B19-B17)/KgVerde)"),
        37: ("PrecioVerdeMaxObj", "=IF(KgVerde=0,0,(B19*(1-MargenObjetivo)-B17)/KgVerde)"),
        38: ("VentaMinCuarto", "=B24"),
        39: ("VentaMinObj", "=IF(1-MargenObjetivo=0,0,B24/(1-MargenObjetivo))"),
        40: ("Factible", '=IF(B21>=MargenObjetivo,"FACTIBLE","REVISAR")'),
        41: ("CostoVerdeCuarto", "=IF(B4=0,0,B9/B4)"),
        42: ("CostoMaquilaCuarto", "=IF(B4=0,0,B10/B4)"),
        43: ("CostoBolsaCuarto", "=IF(B4=0,0,B11/B4)"),
        44: ("CostoEtiqCuarto", "=IF(B4=0,0,B12/B4)"),
        45: ("CostoOtrosCuarto", "=IF(B4=0,0,(B13+B14+B15)/B4)"),
        46: ("Rendimiento", "=1-MermaTostado"),
        47: ("CuartosPorKgVerde", "=IF(KgVerde=0,0,B4/KgVerde)"),
    }
    ws_c["A1"] = "Clave"
    ws_c["B1"] = "Valor"
    for r, (k, f) in formulas.items():
        ws_c.cell(r, 1, k)
        ws_c.cell(r, 2, f)
    ws_c.sheet_state = "hidden"

    calc_names = {
        "KgTostado": "Calc!$B$2",
        "KgPorCuarto": "Calc!$B$3",
        "Cuartos": "Calc!$B$4",
        "SobranteG": "Calc!$B$5",
        "Bolsas": "Calc!$B$6",
        "KgMaquila": "Calc!$B$7",
        "CostoVerde": "Calc!$B$9",
        "CostoMaquila": "Calc!$B$10",
        "CostoBolsas": "Calc!$B$11",
        "CostoEtiquetas": "Calc!$B$12",
        "CostoMO": "Calc!$B$13",
        "CostoFlete": "Calc!$B$14",
        "CostoOtros": "Calc!$B$15",
        "CostoTotal": "Calc!$B$16",
        "CostoSinVerde": "Calc!$B$17",
        "VentaNeta": "Calc!$B$19",
        "Utilidad": "Calc!$B$20",
        "Margen": "Calc!$B$21",
        "Markup": "Calc!$B$22",
        "CostoCuarto": "Calc!$B$24",
        "CostoKgTostado": "Calc!$B$25",
        "UtilidadCuarto": "Calc!$B$26",
        "PrecioKgTostadoVenta": "Calc!$B$27",
        "IvaCredito": "Calc!$B$29",
        "IvaDebito": "Calc!$B$30",
        "IvaPagar": "Calc!$B$31",
        "CajaSalida": "Calc!$B$32",
        "CajaEntrada": "Calc!$B$33",
        "FlujoCaja": "Calc!$B$34",
        "PrecioVerdeMaxCero": "Calc!$B$36",
        "PrecioVerdeMaxObj": "Calc!$B$37",
        "VentaMinCuarto": "Calc!$B$38",
        "VentaMinObj": "Calc!$B$39",
        "Factible": "Calc!$B$40",
        "CostoVerdeCuarto": "Calc!$B$41",
        "CostoMaquilaCuarto": "Calc!$B$42",
        "CostoBolsaCuarto": "Calc!$B$43",
        "CostoEtiqCuarto": "Calc!$B$44",
        "CostoOtrosCuarto": "Calc!$B$45",
        "Rendimiento": "Calc!$B$46",
        "CuartosPorKgVerde": "Calc!$B$47",
    }
    for n, ref in calc_names.items():
        wb.defined_names.add(DefinedName(name=n, attr_text=ref))

    # ------------------------------------------------------------------
    # HOJA RESUMEN
    # ------------------------------------------------------------------
    ws_r = wb.create_sheet("Resumen", 0)
    COLS_R = 8
    apply_widths(ws_r, [24, 18, 18, 18, 18, 18, 18, 22])
    header_bar(
        ws_r,
        1,
        COLS_R,
        "¿Conviene hacer cuartos para Tostado Club?",
        "Caso base con café verde a $9.990 + IVA/kg  ·  Cambia supuestos en la hoja Parámetros",
    )
    ws_r.row_dimensions[1].height = 34
    ws_r.row_dimensions[2].height = 22

    # Veredicto grande
    ws_r.merge_cells("A4:C4")
    ws_r["A4"] = "VEREDICTO (caso base)"
    ws_r["A4"].font = font_white_b
    ws_r["A4"].fill = fill_cafe
    ws_r["A4"].alignment = center
    for c in range(1, 4):
        ws_r.cell(4, c).fill = fill_cafe
        ws_r.cell(4, c).border = thin

    ws_r.merge_cells("A5:C6")
    ws_r["A5"] = '=Factible'
    ws_r["A5"].font = Font(name="Calibri", size=28, bold=True, color=CAFE)
    ws_r["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_r["A5"].border = thick_gold
    ws_r.row_dimensions[5].height = 28
    ws_r.row_dimensions[6].height = 28

    ws_r.merge_cells("A7:C7")
    ws_r["A7"] = (
        '=IF(Factible="FACTIBLE",'
        '"El margen neto supera el mínimo de "&TEXT(MargenObjetivo,"0%")&" definido en Parámetros.",'
        '"El margen queda bajo el mínimo de "&TEXT(MargenObjetivo,"0%")&". Sube precio, baja verde/maquila o empaque.")'
    )
    ws_r["A7"].font = font_small
    ws_r["A7"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_r.row_dimensions[7].height = 32

    ws_r.conditional_formatting.add(
        "A5:C6",
        FormulaRule(formula=['Factible="FACTIBLE"'], fill=fill_verde,
                    font=Font(name="Calibri", size=28, bold=True, color=VERDE)),
    )
    ws_r.conditional_formatting.add(
        "A5:C6",
        FormulaRule(formula=['Factible="REVISAR"'], fill=fill_rojo,
                    font=Font(name="Calibri", size=28, bold=True, color=ROJO)),
    )

    # KPIs
    kpis = [
        (4, 5, "Cuartos producidos", '=Cuartos', N0),
        (4, 7, "Kg tostados", '=KgTostado', N1),
        (8, 1, "Costo por cuarto (neto)", '=CostoCuarto', CLP_D),
        (8, 3, "Venta por cuarto (neto)", '=PrecioVentaCuarto', CLP),
        (8, 5, "Utilidad por cuarto", '=UtilidadCuarto', CLP_D),
        (8, 7, "Margen neto", '=Margen', PCT),
        (11, 1, "Inversión neta del lote", '=CostoTotal', CLP),
        (11, 3, "Caja a desembolsar (c/IVA)", '=CajaSalida', CLP),
        (11, 5, "Venta neta del lote", '=VentaNeta', CLP),
        (11, 7, "Utilidad neta del lote", '=Utilidad', CLP),
    ]

    def kpi_block(ws, r, c, title, formula, fmt, span=2):
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
        t = ws.cell(r, c, title)
        t.font = font_white
        t.fill = fill_med
        t.alignment = center
        t.border = thin
        ws.cell(r, c + span - 1).fill = fill_med
        ws.cell(r, c + span - 1).border = thin
        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 2, end_column=c + span - 1)
        v = ws.cell(r + 1, c, formula)
        v.font = font_kpi
        v.fill = fill_crema
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.number_format = fmt
        v.border = thin
        ws.cell(r + 1, c + span - 1).border = thin
        ws.cell(r + 2, c).border = thin
        ws.cell(r + 2, c + span - 1).border = thin
        for rr in (r + 1, r + 2):
            for cc in range(c, c + span):
                ws.cell(rr, cc).fill = fill_crema
                ws.cell(rr, cc).border = thin

    for r, c, title, formula, fmt in kpis:
        kpi_block(ws_r, r, c, title, formula, fmt)

    ws_r.row_dimensions[9].height = 22
    ws_r.row_dimensions[10].height = 22
    ws_r.row_dimensions[12].height = 22
    ws_r.row_dimensions[13].height = 22

    # Mini tabla escenarios
    section(ws_r, 15, COLS_R, "Atajo: los 5 precios de café verde (detalle en hoja Escenarios)")

    headers_esc = ["", "A  $9.000", "B  $9.500", "C  $9.990  base", "D  $10.000", "E  $10.500", "", ""]
    for i, h in enumerate(headers_esc, 1):
        cell = ws_r.cell(16, i, h)
        cell.font = font_white_b
        cell.fill = fill_cafe
        cell.alignment = center
        cell.border = thin

    # Point to Escenarios sheet rows we'll define:
    # Escenarios!C10 utilidad, C11 margen, C12 factible — columns C-G for A-E
    ws_r["A17"] = "Utilidad neta lote"
    ws_r["A18"] = "Margen neto"
    ws_r["A19"] = "¿Factible?"
    ws_r["A20"] = "Costo por cuarto"
    for r in range(17, 21):
        ws_r.cell(r, 1).font = font_label
        ws_r.cell(r, 1).fill = fill_crema
        ws_r.cell(r, 1).border = thin
        ws_r.cell(r, 1).alignment = left

    # Escenarios sheet layout planned:
    # Row 8 headers A-E in C-G
    # Row 22 utilidad
    # Row 23 margen
    # Row 24 factible
    # Row 18 costo cuarto
    for col, letter in enumerate(["C", "D", "E", "F", "G"], 2):
        ws_r.cell(17, col, f"=Escenarios!{letter}22")
        style_out(ws_r.cell(17, col), CLP)
        ws_r.cell(18, col, f"=Escenarios!{letter}23")
        style_out(ws_r.cell(18, col), PCT)
        ws_r.cell(19, col, f"=Escenarios!{letter}25")
        style_out(ws_r.cell(19, col))
        ws_r.cell(19, col).alignment = center
        ws_r.cell(20, col, f"=Escenarios!{letter}18")
        style_out(ws_r.cell(20, col), CLP_D)

    ws_r.merge_cells("G16:H16")
    ws_r.merge_cells("G17:H17")
    ws_r.merge_cells("G18:H18")
    ws_r.merge_cells("G19:H19")
    ws_r.merge_cells("G20:H20")

    ws_r.conditional_formatting.add(
        "B18:F18",
        ColorScaleRule(start_type="min", start_color="F8D7DA",
                       mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                       end_type="max", end_color="D8F3E2"),
    )
    ws_r.conditional_formatting.add(
        "B19:F19",
        FormulaRule(formula=['B19="SÍ"'], fill=fill_verde, font=Font(bold=True, color=VERDE)),
    )
    ws_r.conditional_formatting.add(
        "B19:F19",
        FormulaRule(formula=['B19="NO"'], fill=fill_rojo, font=Font(bold=True, color=ROJO)),
    )

    # Umbrales
    section(ws_r, 22, COLS_R, "Umbrales del caso base (con los supuestos actuales)")

    umbrales = [
        (23, "Precio verde máximo para no perder (utilidad = 0)", "=PrecioVerdeMaxCero", CLP, "$/kg neto"),
        (24, "Precio verde máximo para cumplir margen objetivo", "=PrecioVerdeMaxObj", CLP, "$/kg neto"),
        (25, "Precio mínimo de venta por cuarto para no perder", "=VentaMinCuarto", CLP_D, "$/un neto"),
        (26, "Precio mínimo de venta por cuarto para margen objetivo", "=VentaMinObj", CLP_D, "$/un neto"),
        (27, "IVA a pagar (débito − crédito) si se vende todo el lote", "=IvaPagar", CLP, "$"),
        (28, "Flujo de caja del ciclo (cobro − pago, con IVA)", "=FlujoCaja", CLP, "$"),
        (29, "Sobrante de café tostado no empaquetable en cuartos", "=SobranteG", N1, "g"),
    ]
    ws_r["A23"].parent  # noop
    for r, lab, f, fmt, uni in umbrales:
        ws_r.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        label(ws_r, r, 1, lab)
        for c in range(2, 5):
            ws_r.cell(r, c).fill = fill_crema
            ws_r.cell(r, c).border = thin
        ws_r.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws_r.cell(r, 5, f)
        style_out(ws_r.cell(r, 5), fmt, fill_crema2)
        ws_r.cell(r, 6).border = thin
        ws_r.cell(r, 6).fill = fill_crema2
        ws_r.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        ws_r.cell(r, 7, uni)
        ws_r.cell(r, 7).font = font_small
        ws_r.cell(r, 7).alignment = left
        ws_r.cell(r, 7).fill = fill_white
        ws_r.cell(r, 7).border = thin
        ws_r.cell(r, 8).border = thin

    # Lectura comercial
    section(ws_r, 31, COLS_R, "Lectura comercial (contexto)")

    ws_r.merge_cells("A32:H34")
    ws_r["A32"] = (
        "Tostado Club vende al público la suscripción de 500 g (2 cuartos) a $23.990 IVA incl., "
        "es decir ~$11.995 IVA incl. por cuarto (~$10.080 neto). Si te pagan $6.000 + IVA por cuarto, "
        "el cliente se queda con un markup retail típico de club/suscripción y tú con el margen industrial. "
        "Esta planilla NO incluye flete ni mano de obra de empaque salvo que los completes en Parámetros. "
        "Asume que se vende el 100% de los cuartos del lote. "
        "Bolsas y etiquetas son SUPUESTOS: cámbialos apenas tengas cotización real. "
        "La merma de tostado mueve mucho el costo por cuarto: mídela en tu primer lote."
    )
    ws_r["A32"].font = Font(name="Calibri", size=11, color=CAFE)
    ws_r["A32"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws_r["A32"].fill = fill_crema
    for r in range(32, 35):
        for c in range(1, 9):
            ws_r.cell(r, c).fill = fill_crema
            ws_r.cell(r, c).border = thin
    ws_r.row_dimensions[32].height = 22
    ws_r.row_dimensions[33].height = 22
    ws_r.row_dimensions[34].height = 22

    ws_r.merge_cells("A36:H36")
    ws_r["A36"] = "Coffee Roasting Labs  ·  RUT 77.586.349-8  ·  Planilla de trabajo interno  ·  Edita solo celdas amarillas en Parámetros"
    ws_r["A36"].font = font_small
    ws_r["A36"].alignment = center

    ws_r.sheet_properties.tabColor = VERDE
    ws_r.freeze_panes = "A4"
    ws_r.sheet_view.showGridLines = False
    ws_r.page_setup.orientation = "landscape"
    ws_r.page_setup.fitToPage = True
    ws_r.page_setup.fitToWidth = 1
    ws_r.page_setup.fitToHeight = 1
    ws_r.print_area = "A1:H36"

    # ------------------------------------------------------------------
    # HOJA LOTE
    # ------------------------------------------------------------------
    ws_l = wb.create_sheet("Lote 1 tonelada")
    COLS_L = 7
    apply_widths(ws_l, [42, 18, 18, 18, 16, 22, 28])
    header_bar(
        ws_l,
        1,
        COLS_L,
        "Estado de resultados del lote  ·  caso base",
        "Café verde al precio de Parámetros (hoy $9.990 + IVA/kg)  ·  cifras NETAS salvo donde se indica IVA",
    )

    section(ws_l, 4, COLS_L, "A. Conversión de 1 tonelada de café verde a cuartos")

    conv = [
        (5, "Café verde comprado", "=KgVerde", N0, "kg", "Pedido de 1.000 kg (1 tonelada)"),
        (6, "Rendimiento de tostado", "=Rendimiento", PCT, "%", "1 − merma"),
        (7, "Café tostado obtenido", "=KgTostado", N1, "kg", "Verde × rendimiento"),
        (8, "Peso de cada cuarto", "=PesoCuartoG", N0, "g", "Café tostado empacado"),
        (9, "Cuartos enteros producidos", "=Cuartos", N0, "un", "Se redondea hacia abajo: no se vende fracción"),
        (10, "Sobrante tostado", "=SobranteG", N1, "g", "Queda para muestra, merma o reempaque"),
        (11, "Cuartos por kg de verde", "=CuartosPorKgVerde", N2, "un/kg", "Con 18% merma ≈ 3,28 cuartos / kg verde"),
        (12, "Bolsas y etiquetas a comprar", "=Bolsas", N0, "un", "Cuartos + merma de empaque"),
        (13, "Kg sobre los que se cobra maquila", "=KgMaquila", N1, "kg", "Verde o tostado, según Parámetros"),
    ]
    for r, lab, f, fmt, uni, how in conv:
        label(ws_l, r, 1, lab)
        ws_l.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws_l.cell(r, 2, f)
        style_out(ws_l.cell(r, 2), fmt, fill_crema2)
        ws_l.cell(r, 3).border = thin
        ws_l.cell(r, 3).fill = fill_crema2
        ws_l.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws_l.cell(r, 4, uni).alignment = center
        ws_l.cell(r, 4).font = font_small
        ws_l.cell(r, 4).border = thin
        ws_l.cell(r, 5).border = thin
        ws_l.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        note(ws_l, r, 6, how)
        ws_l.cell(r, 6).border = thin
        ws_l.cell(r, 7).border = thin

    section(ws_l, 15, COLS_L, "B. Costos del lote (NETOS, sin IVA)")

    for i, h in enumerate(["Ítem", "Base de cálculo", "Precio unit. neto", "Costo neto lote", "% del costo", "Costo por cuarto", "Costo c/IVA lote"], 1):
        c = ws_l.cell(16, i, h)
        c.font = font_white_b
        c.fill = fill_cafe
        c.alignment = center
        c.border = thin

    cost_rows = [
        (17, "Café verde", '=TEXT(KgVerde,"#,##0")&" kg"', "=PrecioVerdeBase", "=CostoVerde", "=CostoVerdeCuarto"),
        (18, "Maquila (tostado)", '=TEXT(KgMaquila,"#,##0.0")&" kg"', "=MaquilaKg", "=CostoMaquila", "=CostoMaquilaCuarto"),
        (19, "Bolsas", '=TEXT(Bolsas,"#,##0")&" un"', "=PrecioBolsa", "=CostoBolsas", "=CostoBolsaCuarto"),
        (20, "Etiquetas", '=TEXT(Bolsas,"#,##0")&" un"', "=PrecioEtiqueta", "=CostoEtiquetas", "=CostoEtiqCuarto"),
        (21, "Mano de obra empaque", '=TEXT(Cuartos,"#,##0")&" un"', "=MOEmpaqueCuarto", "=CostoMO", "=IF(Cuartos=0,0,CostoMO/Cuartos)"),
        (22, "Flete / logística", "1 lote", "=FleteLote", "=CostoFlete", "=IF(Cuartos=0,0,CostoFlete/Cuartos)"),
        (23, "Otros", "1 lote", "=OtrosLote", "=CostoOtros", "=IF(Cuartos=0,0,CostoOtros/Cuartos)"),
    ]
    for r, lab, base, punit, total, xcuarto in cost_rows:
        label(ws_l, r, 1, lab, fill_white if r % 2 else fill_crema)
        ws_l.cell(r, 2, base).border = thin
        ws_l.cell(r, 2).alignment = center
        ws_l.cell(r, 2).font = font_num
        style_out(ws_l.cell(r, 3), CLP)
        ws_l.cell(r, 3, punit)
        style_out(ws_l.cell(r, 3), CLP)
        ws_l.cell(r, 4, total)
        style_out(ws_l.cell(r, 4), CLP)
        ws_l.cell(r, 5, f"=IF(CostoTotal=0,0,D{r}/CostoTotal)")
        style_out(ws_l.cell(r, 5), PCT)
        ws_l.cell(r, 6, xcuarto)
        style_out(ws_l.cell(r, 6), CLP_D)
        ws_l.cell(r, 7, f"=D{r}*(1+IVA)")
        style_out(ws_l.cell(r, 7), CLP)
        if r % 2 == 0:
            for c in range(1, 8):
                if ws_l.cell(r, c).fill.fgColor.rgb in ("00000000", None, "00FFFFFF"):
                    ws_l.cell(r, c).fill = fill_crema

    # Total
    for c in range(1, 8):
        ws_l.cell(24, c).fill = fill_cafe
        ws_l.cell(24, c).font = font_white_b
        ws_l.cell(24, c).border = thin
    ws_l["A24"] = "TOTAL COSTO LOTE"
    ws_l["D24"] = "=CostoTotal"
    ws_l["D24"].number_format = CLP
    ws_l["D24"].font = font_white_b
    ws_l["E24"] = 1
    ws_l["E24"].number_format = PCT
    ws_l["E24"].font = font_white_b
    ws_l["F24"] = "=CostoCuarto"
    ws_l["F24"].number_format = CLP_D
    ws_l["F24"].font = font_white_b
    ws_l["G24"] = "=CajaSalida"
    ws_l["G24"].number_format = CLP
    ws_l["G24"].font = font_white_b
    ws_l["A24"].font = font_white_b

    section(ws_l, 26, COLS_L, "C. Venta, utilidad e IVA (asume venta del 100% de los cuartos)")

    py = [
        (27, "Cuartos vendidos", "=Cuartos", N0),
        (28, "Precio neto por cuarto", "=PrecioVentaCuarto", CLP),
        (29, "Ingreso neto del lote", "=VentaNeta", CLP),
        (30, "Costo neto del lote", "=CostoTotal", CLP),
        (31, "UTILIDAD NETA DEL LOTE", "=Utilidad", CLP),
        (32, "Margen neto (utilidad / venta)", "=Margen", PCT),
        (33, "Markup (utilidad / costo)", "=Markup", PCT),
        (34, "Costo por kg tostado", "=CostoKgTostado", CLP_D),
        (35, "Precio de venta por kg tostado", "=PrecioKgTostadoVenta", CLP),
        (36, "IVA crédito (compras)", "=IvaCredito", CLP),
        (37, "IVA débito (ventas)", "=IvaDebito", CLP),
        (38, "IVA a pagar al SII", "=IvaPagar", CLP),
        (39, "Desembolso de caja (costos + IVA)", "=CajaSalida", CLP),
        (40, "Cobro de caja (ventas + IVA)", "=CajaEntrada", CLP),
        (41, "Flujo de caja del ciclo", "=FlujoCaja", CLP),
    ]
    for r, lab, f, fmt in py:
        highlight = r in (31, 32, 41)
        lab_fill = fill_verde if highlight else fill_crema
        label(ws_l, r, 1, lab, lab_fill)
        ws_l.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws_l.cell(r, 2, f)
        style_out(ws_l.cell(r, 2), fmt, fill_verde if highlight else fill_crema2)
        ws_l.cell(r, 2).font = Font(name="Calibri", size=12, bold=True, color=CAFE) if highlight else font_num
        ws_l.cell(r, 3).border = thin
        ws_l.cell(r, 3).fill = fill_verde if highlight else fill_crema2
        ws_l.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        for c in range(4, 8):
            ws_l.cell(r, c).border = thin
            ws_l.cell(r, c).fill = fill_verde if highlight else fill_white

    notes_lote = (
        "El IVA no es ganancia ni pérdida: lo recaudas/pagas al SII. El margen que importa para factibilidad es el NETO. "
        "La caja sí se ve afectada: compras 1 tonelada de una vez (sale IVA) y cobras cuando factures los cuartos (entra IVA). "
        "Si Tostado Club paga a 30 días, hay que financiar el desembolso de caja hasta el cobro."
    )
    ws_l.merge_cells("A43:G44")
    ws_l["A43"] = notes_lote
    ws_l["A43"].font = font_small
    ws_l["A43"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws_l["A43"].fill = fill_crema
    for r in (43, 44):
        for c in range(1, 8):
            ws_l.cell(r, c).fill = fill_crema
            ws_l.cell(r, c).border = thin
    ws_l.row_dimensions[43].height = 20
    ws_l.row_dimensions[44].height = 20

    ws_l.sheet_properties.tabColor = CAFE_MED
    ws_l.freeze_panes = "A5"
    ws_l.sheet_view.showGridLines = False
    ws_l.page_setup.orientation = "landscape"
    ws_l.page_setup.fitToPage = True
    ws_l.page_setup.fitToWidth = 1
    ws_l.page_setup.fitToHeight = 1

    # ------------------------------------------------------------------
    # HOJA COSTO UNITARIO
    # ------------------------------------------------------------------
    ws_u = wb.create_sheet("Costo por cuarto")
    COLS_U = 6
    apply_widths(ws_u, [36, 18, 18, 16, 22, 30])
    header_bar(
        ws_u,
        1,
        COLS_U,
        "Desglose de 1 cuarto de 250 g  ·  caso base",
        "Cada peso del costo se asigna a una bolsa lista para entregar a Tostado Club",
    )

    section(ws_u, 4, COLS_U, "Estructura de costo de un cuarto")

    for i, h in enumerate(["Componente", "$ neto / cuarto", "% del costo", "% del precio", "Gráfico", "Comentario"], 1):
        c = ws_u.cell(5, i, h)
        c.font = font_white_b
        c.fill = fill_cafe
        c.alignment = center
        c.border = thin

    unit_rows = [
        (6, "Café verde", "=CostoVerdeCuarto", "El verde se encarece con la merma: pagas 1 kg y empacas ~820 g."),
        (7, "Maquila", "=CostoMaquilaCuarto", "Tostado. Si la maquila se cobra sobre verde, también la absorbe la merma."),
        (8, "Bolsa", "=CostoBolsaCuarto", "Incluye las bolsas extra de merma de empaque prorrateadas."),
        (9, "Etiqueta", "=CostoEtiqCuarto", "Una etiqueta por bolsa comprada."),
        (10, "MO + flete + otros", "=CostoOtrosCuarto", "Cero hasta que completes esos campos en Parámetros."),
    ]
    for r, lab, f, com in unit_rows:
        label(ws_u, r, 1, lab, fill_white if r % 2 else fill_crema)
        ws_u.cell(r, 2, f)
        style_out(ws_u.cell(r, 2), CLP_D)
        ws_u.cell(r, 3, f"=IF(CostoCuarto=0,0,B{r}/CostoCuarto)")
        style_out(ws_u.cell(r, 3), PCT)
        ws_u.cell(r, 4, f"=IF(PrecioVentaCuarto=0,0,B{r}/PrecioVentaCuarto)")
        style_out(ws_u.cell(r, 4), PCT)
        ws_u.cell(r, 5, f"=C{r}")  # helper for chart
        style_out(ws_u.cell(r, 5), PCT)
        note(ws_u, r, 6, com)
        ws_u.cell(r, 6).border = thin

    for c in range(1, 7):
        ws_u.cell(11, c).fill = fill_cafe
        ws_u.cell(11, c).font = font_white_b
        ws_u.cell(11, c).border = thin
    ws_u["A11"] = "COSTO TOTAL POR CUARTO"
    ws_u["B11"] = "=CostoCuarto"
    ws_u["B11"].number_format = CLP_D
    ws_u["B11"].font = font_white_b
    ws_u["C11"] = 1
    ws_u["C11"].number_format = PCT
    ws_u["C11"].font = font_white_b
    ws_u["D11"] = "=IF(PrecioVentaCuarto=0,0,CostoCuarto/PrecioVentaCuarto)"
    ws_u["D11"].number_format = PCT
    ws_u["D11"].font = font_white_b

    for c in range(1, 7):
        ws_u.cell(12, c).fill = fill_verde_s
        ws_u.cell(12, c).font = font_white_b
        ws_u.cell(12, c).border = thin
    ws_u["A12"] = "PRECIO DE VENTA"
    ws_u["B12"] = "=PrecioVentaCuarto"
    ws_u["B12"].number_format = CLP
    ws_u["B12"].font = font_white_b
    ws_u["D12"] = 1
    ws_u["D12"].number_format = PCT
    ws_u["D12"].font = font_white_b

    for c in range(1, 7):
        ws_u.cell(13, c).fill = fill_gold
        ws_u.cell(13, c).font = Font(name="Calibri", size=11, bold=True, color=CAFE)
        ws_u.cell(13, c).border = thin
    ws_u["A13"] = "UTILIDAD POR CUARTO"
    ws_u["B13"] = "=UtilidadCuarto"
    ws_u["B13"].number_format = CLP_D
    ws_u["B13"].font = Font(name="Calibri", size=12, bold=True, color=CAFE)
    ws_u["D13"] = "=Margen"
    ws_u["D13"].number_format = PCT
    ws_u["D13"].font = Font(name="Calibri", size=12, bold=True, color=CAFE)

    pie_like = BarChart()
    pie_like.type = "bar"
    pie_like.grouping = "stacked"
    pie_like.title = "Composición del costo por cuarto"
    pie_like.y_axis.title = None
    pie_like.x_axis.title = None
    data = Reference(ws_u, min_col=2, min_row=6, max_row=10)
    cats = Reference(ws_u, min_col=1, min_row=6, max_row=10)
    pie_like.add_data(data, from_rows=False, titles_from_data=False)
    pie_like.set_categories(cats)
    pie_like.shape = 4
    pie_like.style = 10
    pie_like.legend = None
    pie_like.y_axis.numFmt = '"$"#,##0'
    pie_like.width = 18
    pie_like.height = 8
    ws_u.add_chart(pie_like, "A15")

    ws_u.merge_cells("A31:F33")
    ws_u["A31"] = (
        "Regla rápida: con merma 18%, cada kg verde rinde 3,28 cuartos. "
        "Costo verde por cuarto ≈ precio_verde / 3,28. A $9.990 eso es ~$3.046 de verde por bolsa, "
        "antes de maquila y empaque. Si la merma sube a 20%, rinde 3,20 cuartos y el verde por bolsa sube a ~$3.122."
    )
    ws_u["A31"].font = font_small
    ws_u["A31"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws_u["A31"].fill = fill_crema
    for r in range(31, 34):
        for c in range(1, 7):
            ws_u.cell(r, c).fill = fill_crema
            ws_u.cell(r, c).border = thin

    ws_u.sheet_properties.tabColor = DORADO
    ws_u.sheet_view.showGridLines = False
    ws_u.page_setup.orientation = "landscape"
    ws_u.page_setup.fitToPage = True
    ws_u.page_setup.fitToWidth = 1
    ws_u.page_setup.fitToHeight = 1

    # ------------------------------------------------------------------
    # HOJA ESCENARIOS
    # ------------------------------------------------------------------
    ws_e = wb.create_sheet("Escenarios")
    COLS_E = 8
    apply_widths(ws_e, [38, 16, 16, 16, 16, 16, 16, 22])
    header_bar(
        ws_e,
        1,
        COLS_E,
        "Escenarios de precio del café verde",
        "Misma tonelada, misma maquila, mismas bolsas/etiquetas. Solo cambia el $/kg neto del verde.",
    )

    ws_e.merge_cells("A4:H4")
    ws_e["A4"] = (
        "Los precios de cada columna se editan en Parámetros filas 23–27. "
        "El resto (merma, maquila, empaque, venta) se toma de Parámetros."
    )
    ws_e["A4"].font = font_small
    ws_e["A4"].fill = fill_gold
    ws_e["A4"].alignment = Alignment(indent=1, vertical="center")
    fill_row(ws_e, 4, COLS_E, fill_gold)

    # Column headers C-G = scenarios A-E. Column B = base reference unused.
    # Layout:
    # Row 6: scenario names
    # Row 7: green price (linked)
    headers_e = ["Indicador", "Unidad", "A", "B", "C  base", "D", "E", ""]
    for i, h in enumerate(headers_e, 1):
        c = ws_e.cell(6, i, h)
        c.font = font_white_b
        c.fill = fill_cafe
        c.alignment = center
        c.border = thin

    ws_e["A7"] = "Precio café verde"
    ws_e["B7"] = "$/kg neto"
    ws_e["C7"] = "=EscA"
    ws_e["D7"] = "=EscB"
    ws_e["E7"] = "=EscC"
    ws_e["F7"] = "=EscD"
    ws_e["G7"] = "=EscE"
    for col in range(3, 8):
        style_input(ws_e.cell(7, col), CLP)
        # These are outputs of named ranges, still gold so visible; they should NOT be edited here
        ws_e.cell(7, col).fill = fill_crema2
        ws_e.cell(7, col).border = thin
        ws_e.cell(7, col).font = Font(name="Calibri", size=12, bold=True, color=CAFE)

    label(ws_e, 7, 1, "Precio café verde")
    ws_e["B7"].fill = fill_crema
    ws_e["B7"].border = thin
    ws_e["B7"].alignment = center
    ws_e["B7"].font = font_small
    ws_e["H7"] = "Editar en Parámetros"
    ws_e["H7"].font = font_small

    # Helper formulas per column. Green price is row 7.
    # Kg verde, merma etc. are shared.
    #
    # Costo verde col = KgVerde * C7
    # Costo maquila = KgMaquila * MaquilaKg   (independent of green)
    # packing independent
    # Costo total = verde + CostoSinVerde  but CostoSinVerde is from BASE calc which is correct
    #   (doesn't depend on green price). Yes!
    #
    # So for each scenario:
    # costo_verde = KgVerde * precio
    # costo_total = costo_verde + CostoSinVerde
    # venta = VentaNeta (same)
    # utilidad = venta - costo_total
    # margen = utilidad/venta
    # costo_cuarto = costo_total/Cuartos
    # etc.

    def esc_row(r, lab, unidad, formula_c, fmt, fill=None):
        label(ws_e, r, 1, lab, fill or fill_crema)
        ws_e.cell(r, 2, unidad).font = font_small
        ws_e.cell(r, 2).alignment = center
        ws_e.cell(r, 2).border = thin
        ws_e.cell(r, 2).fill = fill or fill_crema
        # formula_c is a template with {c} for column letter and {r7} conceptually using C$7 etc
        for i, col in enumerate(["C", "D", "E", "F", "G"]):
            f = formula_c.replace("{col}", col)
            ws_e.cell(r, i + 3, f)
            style_out(ws_e.cell(r, i + 3), fmt, fill or fill_white)
        ws_e.cell(r, 8).border = thin

    esc_row(8, "Costo café verde", "$ neto", "=KgVerde*{col}7", CLP)
    esc_row(9, "Maquila + empaque + otros", "$ neto", "=CostoSinVerde", CLP)
    esc_row(10, "Costo total lote", "$ neto", "={col}8+{col}9", CLP)
    esc_row(11, "Costo total lote c/IVA", "$", "={col}10*(1+IVA)", CLP)
    esc_row(12, "Kg tostados", "kg", "=KgTostado", N1)
    esc_row(13, "Cuartos producidos", "un", "=Cuartos", N0)
    esc_row(14, "Ingreso neto (venta 100%)", "$ neto", "=VentaNeta", CLP)
    esc_row(15, "Ingreso c/IVA", "$", "=CajaEntrada", CLP)

    esc_row(17, "Costo por kg tostado", "$/kg neto", "=IF(KgTostado=0,0,{col}10/KgTostado)", CLP_D)
    esc_row(18, "Costo por cuarto", "$/un neto", "=IF(Cuartos=0,0,{col}10/Cuartos)", CLP_D)
    esc_row(19, "Utilidad por cuarto", "$/un neto", "=PrecioVentaCuarto-{col}18", CLP_D)
    esc_row(20, "Café verde por cuarto", "$/un neto", "=IF(Cuartos=0,0,{col}8/Cuartos)", CLP_D)

    esc_row(22, "UTILIDAD NETA DEL LOTE", "$ neto", "={col}14-{col}10", CLP, fill_verde)
    esc_row(23, "Margen neto", "%", "=IF({col}14=0,0,{col}22/{col}14)", PCT, fill_verde)
    esc_row(24, "Markup sobre costo", "%", "=IF({col}10=0,0,{col}22/{col}10)", PCT)
    esc_row(25, "¿Cumple margen objetivo?", "SÍ / NO",
            '=IF({col}23>=MargenObjetivo,"SÍ","NO")', None, fill_crema2)
    for col in range(3, 8):
        ws_e.cell(25, col).alignment = center
        ws_e.cell(25, col).font = Font(name="Calibri", size=12, bold=True)

    esc_row(27, "IVA crédito", "$", "={col}10*IVA", CLP)
    esc_row(28, "IVA débito", "$", "=IvaDebito", CLP)
    esc_row(29, "IVA a pagar", "$", "={col}28-{col}27", CLP)
    esc_row(30, "Flujo de caja del ciclo", "$", "={col}15-{col}11", CLP, fill_azul)
    esc_row(31, "Utilidad neta / kg verde", "$/kg", "=IF(KgVerde=0,0,{col}22/KgVerde)", CLP_D)
    esc_row(32, "ROI sobre costo neto", "%", "=IF({col}10=0,0,{col}22/{col}10)", PCT)

    ws_e.conditional_formatting.add(
        "C23:G23",
        ColorScaleRule(start_type="min", start_color="F8D7DA",
                       mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                       end_type="max", end_color="A9DFBF"),
    )
    ws_e.conditional_formatting.add(
        "C25:G25",
        FormulaRule(formula=['C25="SÍ"'], fill=fill_verde, font=Font(bold=True, color=VERDE, size=12)),
    )
    ws_e.conditional_formatting.add(
        "C25:G25",
        FormulaRule(formula=['C25="NO"'], fill=fill_rojo, font=Font(bold=True, color=ROJO, size=12)),
    )
    ws_e.conditional_formatting.add(
        "C22:G22",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill_rojo, font=Font(bold=True, color=ROJO)),
    )

    # Chart data (utilidad) — already in row 22
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Utilidad neta del lote según precio del verde"
    chart.y_axis.title = "Pesos netos"
    chart.y_axis.numFmt = '"$"#,##0'
    data = Reference(ws_e, min_col=3, min_row=21, max_col=7, max_row=22)
    # better: titles from row 6, data row 22
    chart2_data = Reference(ws_e, min_col=2, min_row=22, max_col=7)
    cats = Reference(ws_e, min_col=3, min_row=6, max_col=7)
    # add_data with from_rows
    chart.add_data(Reference(ws_e, min_col=3, min_row=22, max_col=7, max_row=22), from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.legend = None
    chart.width = 18
    chart.height = 8
    ws_e.add_chart(chart, "A34")

    chart_m = BarChart()
    chart_m.type = "col"
    chart_m.title = "Margen neto según precio del verde"
    chart_m.y_axis.title = "Margen"
    chart_m.y_axis.numFmt = "0%"
    chart_m.add_data(Reference(ws_e, min_col=3, min_row=23, max_col=7, max_row=23), from_rows=True, titles_from_data=False)
    chart_m.set_categories(cats)
    chart_m.legend = None
    chart_m.style = 12
    chart_m.width = 18
    chart_m.height = 8
    ws_e.add_chart(chart_m, "A50")

    ws_e.merge_cells("A66:H67")
    ws_e["A66"] = (
        "Cómo leerlo: si todos los escenarios salen SÍ, el negocio aguanta bien una subida del verde hasta $10.500. "
        "Si alguno sale NO, ese precio ya no cumple tu margen mínimo de Parámetros. "
        "Compara también el desembolso de caja (fila 11): es lo que hay que tener disponible para comprar la tonelada, tostar y empacar."
    )
    ws_e["A66"].font = font_small
    ws_e["A66"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws_e["A66"].fill = fill_crema
    for r in (66, 67):
        for c in range(1, 9):
            ws_e.cell(r, c).fill = fill_crema
            ws_e.cell(r, c).border = thin

    ws_e.sheet_properties.tabColor = AZUL
    ws_e.freeze_panes = "C7"
    ws_e.sheet_view.showGridLines = False
    ws_e.page_setup.orientation = "landscape"
    ws_e.page_setup.fitToPage = True
    ws_e.page_setup.fitToWidth = 1
    ws_e.page_setup.fitToHeight = 1

    # ------------------------------------------------------------------
    # HOJA SENSIBILIDAD
    # ------------------------------------------------------------------
    ws_s = wb.create_sheet("Sensibilidad")
    COLS_S = 10
    apply_widths(ws_s, [22, 14, 14, 14, 14, 14, 14, 14, 14, 18])
    header_bar(
        ws_s,
        1,
        COLS_S,
        "Sensibilidad extra  ·  merma, bolsas y precio de venta",
        "Tablas que se recalculan con Parámetros. Útiles si aún no cierras cotización de empaque o el perfil de tueste.",
    )

    # --- Tabla 1: verde vs merma ---
    section(ws_s, 4, COLS_S, "1. Utilidad neta del lote: precio verde (filas) × merma de tostado (columnas)")

    ws_s["A5"] = "Verde \\ Merma"
    ws_s["A5"].font = font_white_b
    ws_s["A5"].fill = fill_cafe
    ws_s["A5"].alignment = center
    ws_s["A5"].border = thin

    mermas = [0.14, 0.16, 0.18, 0.20, 0.22]
    for i, m in enumerate(mermas):
        cell = ws_s.cell(5, i + 2, m)
        cell.number_format = PCT
        cell.font = font_white_b
        cell.fill = fill_cafe
        cell.alignment = center
        cell.border = thin
    ws_s["G5"] = "Lectura"
    ws_s["G5"].font = font_white_b
    ws_s["G5"].fill = fill_cafe
    ws_s["G5"].alignment = center
    ws_s.merge_cells("G5:J5")
    for c in range(8, 11):
        ws_s.cell(5, c).fill = fill_cafe
        ws_s.cell(5, c).border = thin

    # Utility function in Excel for a given green price and merma:
    # kg_tostado = KgVerde * (1-merma)
    # cuartos = INT(kg_tostado / KgPorCuarto)
    # kg_maquila = IF(MaquilaSobre="Tostado", kg_tostado, KgVerde)
    # bolsas = ROUNDUP(cuartos*(1+MermaEmpaque),0)
    # costo = KgVerde*precio + kg_maquila*MaquilaKg + bolsas*(PrecioBolsa+PrecioEtiqueta) + cuartos*MO + Flete + Otros
    # venta = cuartos * PrecioVentaCuarto
    # utilidad = venta - costo
    #
    # We'll put green prices in A6:A10 from EscA-EscE
    # merma headers in B5:F5
    #
    # Formula for B6 (green A6, merma B$5):

    def utilidad_formula(precio_cell, merma_cell):
        # precio_cell like $A6, merma_cell like B$5
        return (
            f'=LET('
            f'kgv,KgVerde,'
            f'mer,{merma_cell},'
            f'kgt,kgv*(1-mer),'
            f'kpc,PesoCuartoG/1000,'
            f'cu,IF(kpc=0,0,INT(kgt/kpc)),'
            f'bol,ROUNDUP(cu*(1+MermaEmpaque),0),'
            f'kgm,IF(MaquilaSobre="Tostado",kgt,kgv),'
            f'costo,kgv*{precio_cell}+kgm*MaquilaKg+bol*(PrecioBolsa+PrecioEtiqueta)+cu*MOEmpaqueCuarto+FleteLote+OtrosLote,'
            f'venta,cu*PrecioVentaCuarto,'
            f'venta-costo)'
        )

    # LET might not work on older Excel. Erik likely has modern Excel/Google? 
    # Google Sheets also has LET. To be safer, expand the formula without LET.
    def utilidad_formula_compat(precio_cell, merma_cell):
        kgt = f"(KgVerde*(1-{merma_cell}))"
        kpc = "(PesoCuartoG/1000)"
        cu = f"IF({kpc}=0,0,INT({kgt}/{kpc}))"
        bol = f"ROUNDUP(({cu})*(1+MermaEmpaque),0)"
        kgm = f'IF(MaquilaSobre="Tostado",{kgt},KgVerde)'
        costo = (
            f"(KgVerde*{precio_cell}+({kgm})*MaquilaKg+"
            f"({bol})*(PrecioBolsa+PrecioEtiqueta)+"
            f"({cu})*MOEmpaqueCuarto+FleteLote+OtrosLote)"
        )
        venta = f"(({cu})*PrecioVentaCuarto)"
        return f"={venta}-{costo}"

    ws_s["A6"] = "=EscA"
    ws_s["A7"] = "=EscB"
    ws_s["A8"] = "=EscC"
    ws_s["A9"] = "=EscD"
    ws_s["A10"] = "=EscE"
    for r in range(6, 11):
        style_out(ws_s.cell(r, 1), CLP, fill_crema)
        ws_s.cell(r, 1).font = font_input
        for c in range(2, 7):
            merma_ref = f"{get_column_letter(c)}$5"
            precio_ref = f"$A{r}"
            ws_s.cell(r, c, utilidad_formula_compat(precio_ref, merma_ref))
            style_out(ws_s.cell(r, c), CLP)
        ws_s.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)

    ws_s["G6"] = "Más merma = menos cuartos = menos venta y el verde se reparte en menos bolsas."
    ws_s["G6"].font = font_small
    ws_s["G6"].alignment = Alignment(wrap_text=True, vertical="center")
    for r in range(6, 11):
        for c in range(7, 11):
            ws_s.cell(r, c).border = thin
            ws_s.cell(r, c).fill = fill_crema

    ws_s.conditional_formatting.add(
        "B6:F10",
        ColorScaleRule(start_type="min", start_color="F5B7B1",
                       mid_type="percentile", mid_value=50, mid_color="F9E79F",
                       end_type="max", end_color="82E0AA"),
    )
    ws_s.conditional_formatting.add(
        "B6:F10",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill_rojo, font=Font(color=ROJO, bold=True)),
    )

    # Highlight default merma 18% column D (index 4 = 0.18 which is 3rd merma = col 4)
    # mermas: 14,16,18,20,22 -> cols B C D E F. 18% is D.

    # --- Tabla 2: verde vs costo bolsa ---
    section(ws_s, 12, COLS_S, "2. Utilidad neta del lote: precio verde × costo de bolsa neta (la etiqueta se mantiene)")

    ws_s["A13"] = "Verde \\ Bolsa"
    ws_s["A13"].font = font_white_b
    ws_s["A13"].fill = fill_cafe
    ws_s["A13"].alignment = center
    ws_s["A13"].border = thin

    bolsas_p = [0, 200, 300, 500, 700]
    for i, b in enumerate(bolsas_p):
        cell = ws_s.cell(13, i + 2, b)
        cell.number_format = CLP
        cell.font = font_white_b
        cell.fill = fill_cafe
        cell.alignment = center
        cell.border = thin
    ws_s.merge_cells("G13:J13")
    ws_s["G13"] = "0 = las pone Tostado Club. $700 ≈ Cafestore. $300 ≈ Funsmart a volumen."
    ws_s["G13"].font = font_white
    ws_s["G13"].fill = fill_cafe
    ws_s["G13"].alignment = left
    for c in range(8, 11):
        ws_s.cell(13, c).fill = fill_cafe
        ws_s.cell(13, c).border = thin

    def utilidad_bolsa(precio_cell, bolsa_cell):
        kgt = "(KgVerde*(1-MermaTostado))"
        kpc = "(PesoCuartoG/1000)"
        cu = f"IF({kpc}=0,0,INT({kgt}/{kpc}))"
        bol = f"ROUNDUP(({cu})*(1+MermaEmpaque),0)"
        kgm = f'IF(MaquilaSobre="Tostado",{kgt},KgVerde)'
        costo = (
            f"(KgVerde*{precio_cell}+({kgm})*MaquilaKg+"
            f"({bol})*({bolsa_cell}+PrecioEtiqueta)+"
            f"({cu})*MOEmpaqueCuarto+FleteLote+OtrosLote)"
        )
        venta = f"(({cu})*PrecioVentaCuarto)"
        return f"={venta}-{costo}"

    ws_s["A14"] = "=EscA"
    ws_s["A15"] = "=EscB"
    ws_s["A16"] = "=EscC"
    ws_s["A17"] = "=EscD"
    ws_s["A18"] = "=EscE"
    for r in range(14, 19):
        style_out(ws_s.cell(r, 1), CLP, fill_crema)
        ws_s.cell(r, 1).font = font_input
        for c in range(2, 7):
            ws_s.cell(r, c, utilidad_bolsa(f"$A{r}", f"{get_column_letter(c)}$13"))
            style_out(ws_s.cell(r, c), CLP)
        ws_s.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
        for c in range(7, 11):
            ws_s.cell(r, c).border = thin
            ws_s.cell(r, c).fill = fill_crema

    ws_s.conditional_formatting.add(
        "B14:F18",
        ColorScaleRule(start_type="min", start_color="F5B7B1",
                       mid_type="percentile", mid_value=50, mid_color="F9E79F",
                       end_type="max", end_color="82E0AA"),
    )

    # --- Tabla 3: precio venta vs verde ---
    section(ws_s, 20, COLS_S, "3. Margen neto: precio verde × precio de venta del cuarto (si negocias otro precio con Tostado Club)")

    ws_s["A21"] = "Verde \\ Venta"
    ws_s["A21"].font = font_white_b
    ws_s["A21"].fill = fill_cafe
    ws_s["A21"].alignment = center
    ws_s["A21"].border = thin

    ventas = [5000, 5500, 6000, 6500, 7000]
    for i, v in enumerate(ventas):
        cell = ws_s.cell(21, i + 2, v)
        cell.number_format = CLP
        cell.font = font_white_b
        cell.fill = fill_cafe
        cell.alignment = center
        cell.border = thin
    ws_s.merge_cells("G21:J21")
    ws_s["G21"] = "Hoy estás en $6.000. Útil si pides alza o te bajan el precio."
    ws_s["G21"].font = font_white
    ws_s["G21"].fill = fill_cafe
    for c in range(8, 11):
        ws_s.cell(21, c).fill = fill_cafe
        ws_s.cell(21, c).border = thin

    def margen_venta(precio_cell, venta_cell):
        kgt = "(KgVerde*(1-MermaTostado))"
        kpc = "(PesoCuartoG/1000)"
        cu = f"IF({kpc}=0,0,INT({kgt}/{kpc}))"
        bol = f"ROUNDUP(({cu})*(1+MermaEmpaque),0)"
        kgm = f'IF(MaquilaSobre="Tostado",{kgt},KgVerde)'
        costo = (
            f"(KgVerde*{precio_cell}+({kgm})*MaquilaKg+"
            f"({bol})*(PrecioBolsa+PrecioEtiqueta)+"
            f"({cu})*MOEmpaqueCuarto+FleteLote+OtrosLote)"
        )
        venta = f"(({cu})*{venta_cell})"
        return f'=IF({venta}=0,0,({venta}-{costo})/{venta})'

    ws_s["A22"] = "=EscA"
    ws_s["A23"] = "=EscB"
    ws_s["A24"] = "=EscC"
    ws_s["A25"] = "=EscD"
    ws_s["A26"] = "=EscE"
    for r in range(22, 27):
        style_out(ws_s.cell(r, 1), CLP, fill_crema)
        ws_s.cell(r, 1).font = font_input
        for c in range(2, 7):
            ws_s.cell(r, c, margen_venta(f"$A{r}", f"{get_column_letter(c)}$21"))
            style_out(ws_s.cell(r, c), PCT)
        ws_s.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
        for c in range(7, 11):
            ws_s.cell(r, c).border = thin
            ws_s.cell(r, c).fill = fill_crema

    ws_s.conditional_formatting.add(
        "B22:F26",
        ColorScaleRule(start_type="min", start_color="F5B7B1",
                       mid_type="percentile", mid_value=50, mid_color="F9E79F",
                       end_type="max", end_color="82E0AA"),
    )
    ws_s.conditional_formatting.add(
        "B22:F26",
        CellIsRule(operator="lessThan", formula=["MargenObjetivo"], fill=fill_rojo),
    )

    ws_s.merge_cells("A28:J30")
    ws_s["A28"] = (
        "Qué no está modelado (y puede mover el número): merma de calidad / quiebre, molienda, cajas master, "
        "financiamiento si Tostado Club paga a 30/60 días, mermas de inventario, y tu tiempo de coordinación. "
        "Si el primer lote real da otra merma, cámbiala en Parámetros y estas tablas se actualizan."
    )
    ws_s["A28"].font = font_small
    ws_s["A28"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws_s["A28"].fill = fill_crema
    for r in range(28, 31):
        for c in range(1, 11):
            ws_s.cell(r, c).fill = fill_crema
            ws_s.cell(r, c).border = thin

    ws_s.sheet_properties.tabColor = ROJO
    ws_s.sheet_view.showGridLines = False
    ws_s.freeze_panes = "B6"
    ws_s.page_setup.orientation = "landscape"
    ws_s.page_setup.fitToPage = True
    ws_s.page_setup.fitToWidth = 1
    ws_s.page_setup.fitToHeight = 1

    # ------------------------------------------------------------------
    # HOJA CÓMO USAR
    # ------------------------------------------------------------------
    ws_i = wb.create_sheet("Como usar", 0)
    COLS_I = 6
    apply_widths(ws_i, [22, 22, 22, 22, 22, 22])
    header_bar(
        ws_i,
        1,
        COLS_I,
        "Planilla de factibilidad  ·  cuartos para Tostado Club",
        "Coffee Roasting Labs  ·  Tostaduría Erik Berwart Araya EIRL  ·  RUT 77.586.349-8",
    )

    def info_block(ws, header_row, body_end, title, body):
        section(ws, header_row, COLS_I, title)
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=body_end, end_column=COLS_I)
        cell = ws.cell(header_row + 1, 1, body)
        cell.font = Font(name="Calibri", size=11, color=CAFE)
        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        for rr in range(header_row + 1, body_end + 1):
            ws.row_dimensions[rr].height = 20
            for c in range(1, 7):
                ws.cell(rr, c).fill = fill_crema
                ws.cell(rr, c).border = thin

    info_block(
        ws_i, 4, 6, "Para qué sirve",
        "Decidir si conviene comprar 1.000 kg de café verde, tostarlo (maquila), empacarlo en cuartos de 250 g "
        "y vendérselos a Tostado Club a $6.000 + IVA. Compara cinco precios de verde y te dice el margen, "
        "la caja que hay que poner y el precio máximo de verde para no perder.",
    )
    info_block(
        ws_i, 7, 9, "Cómo usarla",
        "1) Abre Parámetros y revisa las celdas amarillas. 2) Completa bolsas y etiquetas con tu cotización real "
        "(hoy hay un supuesto). 3) Mira Resumen: veredicto FACTIBLE / REVISAR. 4) En Escenarios ves $9.000, "
        "$9.500, $9.990, $10.000 y $10.500. 5) Sensibilidad sirve si cambia la merma, el precio de la bolsa "
        "o si negocias otro precio de venta.",
    )
    info_block(
        ws_i, 10, 14, "Datos que pediste (ya cargados)",
        "· Compra: 1 tonelada (1.000 kg) de café verde.\n"
        "· Verde: $9.990 + IVA/kg, y escenarios $9.000 / $9.500 / $10.000 / $10.500.\n"
        "· Maquila: $1.500 + IVA/kg (por defecto sobre kg VERDE; cámbialo a Tostado si el maquilador cobra la salida).\n"
        "· Venta: $6.000 + IVA por cuarto de 250 g.\n"
        "· Bolsas y etiquetas: incluidos, con precio supuesto editable.",
    )
    info_block(
        ws_i, 15, 19, "Supuestos que debes validar",
        "· Merma de tostado 18% (tueste medio). Mídela en el primer batch.\n"
        "· Bolsa $300 + IVA (rango de mercado ~$286 a $700 según proveedor).\n"
        "· Etiqueta $80 + IVA (tiraje de ~3.000 un.).\n"
        "· Se vende el 100% del lote. Flete y mano de obra de empaque en 0 hasta que los cotices.\n"
        "· IVA 19%. Los márgenes se miran NETOS; la caja se mira CON IVA.",
    )
    info_block(
        ws_i, 20, 24, "Hojas",
        "Resumen — veredicto y números grandes.\n"
        "Parametros — único lugar para editar.\n"
        "Lote 1 tonelada — estado de resultados completo + IVA.\n"
        "Costo por cuarto — de dónde sale cada peso de la bolsa.\n"
        "Escenarios — los 5 precios de verde, lado a lado.\n"
        "Sensibilidad — merma, bolsas y precio de venta.",
    )

    ws_i.merge_cells("A26:F27")
    ws_i["A26"] = (
        f"Documento interno de trabajo. No es una cotización al cliente. "
        f"Generado {date.today().isoformat()}. "
        "Cambia las celdas amarillas; no borres nombres definidos ni la hoja Calc (está oculta)."
    )
    ws_i["A26"].font = font_small
    ws_i["A26"].alignment = Alignment(wrap_text=True, indent=1, vertical="center")
    for rr in (26, 27):
        for c in range(1, 7):
            ws_i.cell(rr, c).fill = fill_gris
            ws_i.cell(rr, c).border = thin

    ws_i.sheet_properties.tabColor = CAFE
    ws_i.sheet_view.showGridLines = False
    ws_i.page_setup.orientation = "landscape"
    ws_i.page_setup.fitToPage = True
    ws_i.page_setup.fitToWidth = 1
    ws_i.page_setup.fitToHeight = 1

    # print settings / freeze
    for ws in wb.worksheets:
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.horizontalCentered = True
        ws.oddHeader.right.text = "Confidencial — Coffee Roasting Labs"
        ws.oddFooter.right.text = "Página &P de &N"
        ws.oddFooter.left.text = "Factibilidad cuartos Tostado Club"

    # Order: Como usar, Resumen, Parametros, Lote, Costo, Escenarios, Sensibilidad, Calc
    order = ["Como usar", "Resumen", "Parametros", "Lote 1 tonelada", "Costo por cuarto",
             "Escenarios", "Sensibilidad", "Calc"]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(OUT)
    print("Wrote", OUT)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    build()
